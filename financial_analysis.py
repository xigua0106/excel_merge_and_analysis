#-*- coding:utf-8 -*-
from tkinter import *
import configparser  #配置文件
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
import os
import sys
from openpyxl import load_workbook

def srcselectPath():
    config = configparser.ConfigParser()
    config.read("filepath.ini")
    src = config.get("文件路径","srcfiles")
    path_ = askdirectory(initialdir=src)
    srcPath.set(path_)
    
def dstselectPath():
    config = configparser.ConfigParser()
    config.read("filepath.ini")
    dst = config.get("文件路径","dstfile")
    path_=askopenfilename(title='选择最终要保存的文件', filetypes=[('excel', '*.xlsx'), ('All Files', '*')],initialdir=dst)
    dstPath.set(path_)
def settingPanel():
    """设置处理文件的路径，包括存放各个部门提交的文件存放文件夹以及需要输出的文件名称"""
    settingPanel = Toplevel()
    settingPanel.title("路径设置")
    settingPanel.geometry("240x120+400+300")
    settingPanel['bg'] = 'lightblue'
    Label(settingPanel,bg = 'lightblue', text = "数据来源").grid(row = 0, column = 0)

    config = configparser.ConfigParser()
    config.read("filepath.ini")
    src = config.get("文件路径","srcfiles")
    dst = config.get("文件路径","dstfile")

    srcPath.set(src)
    dstPath.set(dst)
    Entry(settingPanel, textvariable = srcPath).grid(row = 0, column = 1)
    Button(settingPanel,bg = 'lightblue',text = "选择", command = srcselectPath).grid(row = 0, column = 2)
    Label(settingPanel, bg = 'lightblue',text = "目的文件").grid(row = 1, column = 0)
    Entry(settingPanel, textvariable = dstPath).grid(row = 1, column = 1)
    Button(settingPanel,bg = 'lightblue',text = "选择", command = dstselectPath).grid(row = 1, column = 2)
    
 #   Button(settingPanel,bg = 'lightblue',text = "保存", command = saveinifile()).grid(row = 2, column = 1)
 #   Button(settingPanel,bg = 'lightblue',text = "关闭", command = settingPanel.destroy).grid(row = 2, column = 2)
    Button(settingPanel,bg = 'lightblue',text = "保存", command = lambda:[saveinifile(),settingPanel.destroy()] ).grid(row = 2, column = 1)
def saveinifile():
    """生成目标文件和原目录的字典"""
    """如果没有设定，则程序直接从filepath.ini文件中读取路径进行操作"""
    config = configparser.ConfigParser() 
    config['文件路径'] = {}
    filepath = config['文件路径']
    filepath['srcfiles'] = srcPath.get() 
    filepath['dstfile'] = dstPath.get()
 
    """写入后缀为.ini的文件"""
    with open('filepath.ini', 'w') as configfile:
        config.write(configfile)
def input_merge_files():
    """数据合并处理，对各个部门统计的数据进行合并填入最终的报表中"""
    
    """如果程序在合并前没有进行设置，则直接从ini文件中读取路径"""
    config = configparser.ConfigParser()
    config.read("filepath.ini")
    src = config.get("文件路径","srcfiles")
    dst = config.get("文件路径","dstfile")

    if (src != ''):
        print(src)
    if (dst != ''):
        print(dst)   
    
    wb_final = load_workbook(dst)                             #打开最终保存的统计报表
    ws_final = wb_final["各部门汇总（所本部附件打印）"]       #打开最终要合并写入的报表页面

    counted = True
    start_pos = 0
    stop_pos = 0
    for root, dirs, files in os.walk(src):          #遍历当前路径下所有要进行写入的表格名称
        for name in files:
            filepath = os.path.join(root, name)             #获取表格名称：名称 = 路径+文件名
            wb_input = load_workbook(filepath)              #打开当前文件
            ws_input = wb_input["各部门汇总（所本部附件打印）"] #打开当前文件的统计页面
            for i in range(7, ws_input.max_column):                         #从第7列开始，可以根据表格进行调整
                if (ws_input.cell(row=2,column = i).value in depart_list):   #判断当前表格中的部门
                    if(ws_input.cell(row=3,column = i).value in position_list): #判断当前人员的编制
                        for j in range(5, ws_input.max_row):                    #从第5列开始为填写的表格内容
                            if ws_input.cell(row=j, column=i).value!= None:      
                                ws_final.cell(row=j, column=i).value = ws_input.cell(row=j, column=i).value
    wb_final.save(dstPath.get())                                                    #保存统计后的表格

def input_data_analysis():
    """从合并后的表格中提取各个编制的统计数据，然后生成最终的合并报表"""
    config = configparser.ConfigParser()
    config.read("filepath.ini")
    dst = config.get("文件路径","dstfile")


    wb_final = load_workbook(dst,data_only = True)            #打开最终保存的统计报表
    ws_final = wb_final["各部门汇总（所本部附件打印）"]       #打开最终要合并写入的报表页面
    ws_final_research = wb_final["研究所统计汇总"]            #研究所统计汇总页面
    ws_final_electronic = wb_final["电子公司统计汇总"]            #电子公司统计汇总页面
    ws_final_techinic = wb_final["技术公司统计汇总"]            #技术司统计汇总页面
    for i in range(7, ws_final.max_column+1):                         #从第7列开始，可以根据表格进行调整
        if (ws_final.cell(row=3,column = i).value in position_person_count):   #判断当前表格中的内容是否为统计技术

#统计民机系统部人员数据，并分别放入三个统计页面
                if (ws_final.cell(row=2,column = i).value == '民机系统部'):
                    if (ws_final.cell(row=3,column = i).value == '615所人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_research.cell(row = j,column = 5).value = ws_final.cell(row=j, column=i).value
                            ws_final_research.cell(row = j,column = 6).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '电子公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_electronic.cell(row = j,column = 5).value = ws_final.cell(row=j, column=i).value
                            ws_final_electronic.cell(row = j,column = 6).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '技术公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 5).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 6).value = ws_final.cell(row=j, column=i+1).value
#统计重点实验室人员数据，并分别放入三个统计页面                        
                if (ws_final.cell(row=2,column = i).value in ['重点','重点实验室']):
                    if (ws_final.cell(row=3,column = i).value == '615所人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 7).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 8).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '电子公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_electronic.cell(row = j,column = 7).value = ws_final.cell(row=j, column=i).value
                            ws_final_electronic.cell(row = j,column = 8).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '技术公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 7).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 8).value = ws_final.cell(row=j, column=i+1).value
#统计通导人员数据，并分别放入三个统计页面 									
                if (ws_final.cell(row=2,column = i).value in ['通导','通导系统部']):
                    if (ws_final.cell(row=3,column = i).value == '615所人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_research.cell(row = j,column = 9).value = ws_final.cell(row=j, column=i).value
                            ws_final_research.cell(row = j,column = 10).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '电子公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_electronic.cell(row = j,column = 9).value = ws_final.cell(row=j, column=i).value
                            ws_final_electronic.cell(row = j,column = 10).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '技术公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 9).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 10).value = ws_final.cell(row=j, column=i+1).value
#统计装备系统部人员数据，并分别放入三个统计页面 									
                if (ws_final.cell(row=2,column = i).value in ['装备','装备系统部']):
                    if (ws_final.cell(row=3,column = i).value == '615所人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_research.cell(row = j,column = 11).value = ws_final.cell(row=j, column=i).value
                            ws_final_research.cell(row = j,column = 12).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '电子公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_electronic.cell(row = j,column = 11).value = ws_final.cell(row=j, column=i).value
                            ws_final_electronic.cell(row = j,column = 12).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '技术公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 11).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 12).value = ws_final.cell(row=j, column=i+1).value
#统计电子部人员数据，并分别放入三个统计页面
                if (ws_final.cell(row=2,column = i).value in ['电子部','电子']):

                    if (ws_final.cell(row=3,column = i).value == '615所人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_research.cell(row = j,column = 13).value = ws_final.cell(row=j, column=i).value
                            ws_final_research.cell(row = j,column = 14).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '电子公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_electronic.cell(row = j,column = 13).value = ws_final.cell(row=j, column=i).value
                            ws_final_electronic.cell(row = j,column = 14).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '技术公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 13).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 14).value = ws_final.cell(row=j, column=i+1).value
#统计系统部人员数据，并分别放入三个统计页面
                if (ws_final.cell(row=2,column = i).value in ['系统','系统部']):
                    if (ws_final.cell(row=3,column = i).value == '615所人数小计'):

                        for  j in range (5,ws_final.max_row):
                            ws_final_research.cell(row = j,column = 15).value = ws_final.cell(row=j, column=i).value
                            ws_final_research.cell(row = j,column = 16).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '电子公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_electronic.cell(row = j,column = 15).value = ws_final.cell(row=j, column=i).value
                            ws_final_electronic.cell(row = j,column = 16).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '技术公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 15).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 16).value = ws_final.cell(row=j, column=i+1).value
#统计软件部人员数据，并分别放入三个统计页面
                if (ws_final.cell(row=2,column = i).value in ['软件部','软件']):
                    if (ws_final.cell(row=3,column = i).value == '615所人数小计'):

                        for  j in range (5,ws_final.max_row):
                            ws_final_research.cell(row = j,column = 17).value = ws_final.cell(row=j, column=i).value
                            ws_final_research.cell(row = j,column = 18).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '电子公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_electronic.cell(row = j,column = 17).value = ws_final.cell(row=j, column=i).value
                            ws_final_electronic.cell(row = j,column = 18).value = ws_final.cell(row=j, column=i+1).value
                    if (ws_final.cell(row=3,column = i).value == '技术公司人数小计'):
                        for  j in range (5,ws_final.max_row):
                            ws_final_techinic.cell(row = j,column = 17).value = ws_final.cell(row=j, column=i).value
                            ws_final_techinic.cell(row = j,column = 18).value = ws_final.cell(row=j, column=i+1).value
        if (ws_final.cell(row=3,column = i).value in ['615所分摊总计']):

            for  j in range (5,ws_final.max_row):
                ws_final_research.cell(row = j,column = 19).value = ws_final.cell(row=j, column=i).value
        if (ws_final.cell(row=3,column = i).value in ['航电公司分摊总计']):

            for  j in range (5,ws_final.max_row):
                ws_final_electronic.cell(row = j,column = 19).value = ws_final.cell(row=j, column=i).value
        if (ws_final.cell(row=3,column = i).value in ['技术公司分摊总计']):

            for  j in range (5,ws_final.max_row):
                ws_final_techinic.cell(row = j,column = 19).value = ws_final.cell(row=j, column=i).value
    wb_final.save(dst)

"""主显示页面"""
root = Tk()
dstPath = StringVar()                                                                            #要存储的最终合并表格
srcPath = StringVar()                                                                            #存放各个部门提交表格的目录
root.title("excel数据处理")
root.geometry("300x80+200+100")
root['bg'] = 'lightblue'


#depart_list = ['民机系统部','民机']    #测试部门名称 可以添加
depart_list = ['民机系统部','民机','重点实验室','重点','通导','通导系统部','装备','装备系统部']    #部门名称 可以添加
position_list = ['615所','技术公司','研究所','电子公司','上海航空电子公司']                        #编制名称 可以添加修改
position_person_count = ['615所人数小计','电子公司人数小计','技术公司人数小计']                    #统计各种编制人数需要
position_money_count = ['615所经费小计','电子公司经费小计','技术公司经费小计']                     #统计各种编制经费需要
Label(root,bg = 'lightblue', text = "    ").grid(row = 0, column = 1)
Label(root,bg = 'lightblue', text = "    ").grid(row = 1, column = 0)
bttn_merge = Button(root,bg = 'lightblue',text = "数据合并",command = input_merge_files).grid(row = 1, column = 1)
Label(root,bg = 'lightblue', text = "    ").grid(row = 1, column = 2)
bttn_process = Button(root,bg = 'lightblue',text = "数据处理",command = input_data_analysis).grid(row = 1, column = 3)
Label(root,bg = 'lightblue', text = "   ").grid(row = 1, column = 4)
Button(root,bg = 'lightblue',text = "设  置",command=settingPanel).grid(row = 1, column = 5)
Label(root,bg = 'lightblue', text = "    ").grid(row = 1, column = 6)
Button(root,bg = 'lightblue',text = "退  出",command = quit).grid(row = 1, column = 7)

root.mainloop()